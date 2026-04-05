from __future__ import annotations


import json
import re
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_FILE = DATA_DIR / "police_enforcement_2024_positive_drug_tests.xlsx"
OUTPUT_RECORDS = DATA_DIR / "police_enforcement_2024_positive_drug_tests.cleaned.json"
OUTPUT_SUMMARY = DATA_DIR / "police_enforcement_2024_positive_drug_tests.summary.json"
OUTPUT_CSV = DATA_DIR / "police_enforcement_2024_positive_drug_tests.cleaned.csv"


def clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"\s+", " ", str(value)).strip()
    return normalized or None


def clean_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def date_to_iso(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def build_records(ws) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    header = [
        clean_text(cell.value) or f"column_{index + 1}"
        for index, cell in enumerate(
            next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        )
    ]

    required_columns = [
        "YEAR",
        "START_DATE",
        "END_DATE",
        "JURISDICTION",
        "LOCATION",
        "AGE_GROUP",
        "METRIC",
        "COUNT",
    ]
    missing_columns = [column for column in required_columns if column not in header]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    col_map = {name: idx for idx, name in enumerate(header)}

    records: list[dict[str, Any]] = []
    quality = {
        "rows_seen": 0,
        "rows_emitted": 0,
        "rows_skipped": 0,
        "missing_required_values": defaultdict(int),
    }

    for row_index, row_cells in enumerate(
        ws.iter_rows(min_row=2, values_only=False), start=2
    ):
        quality["rows_seen"] += 1
        values = [c.value for c in row_cells]

        raw_year = values[col_map["YEAR"]] if "YEAR" in col_map else None
        raw_start = values[col_map["START_DATE"]] if "START_DATE" in col_map else None
        raw_end = values[col_map["END_DATE"]] if "END_DATE" in col_map else None
        jurisdiction = (
            clean_text(values[col_map["JURISDICTION"]])
            if "JURISDICTION" in col_map
            else None
        )
        location = (
            clean_text(values[col_map["LOCATION"]]) if "LOCATION" in col_map else None
        )
        age_group = (
            clean_text(values[col_map["AGE_GROUP"]]) if "AGE_GROUP" in col_map else None
        )
        metric = clean_text(values[col_map["METRIC"]]) if "METRIC" in col_map else None

        period_start = date_to_iso(raw_start)
        period_end = date_to_iso(raw_end)

        required_values = {
            "YEAR": raw_year,
            "START_DATE": period_start,
            "END_DATE": period_end,
            "JURISDICTION": jurisdiction,
            "LOCATION": location,
            "AGE_GROUP": age_group,
            "METRIC": metric,
        }
        failed = False
        for column, value in required_values.items():
            if value is None:
                quality["missing_required_values"][column] += 1
                failed = True
        if failed:
            quality["rows_skipped"] += 1
            continue

        def _raw(key: str) -> str | None:
            idx = col_map.get(key)
            if idx is None or idx >= len(values):
                return None
            return clean_text(values[idx])

        series_year = clean_int(raw_year)
        count = clean_int(_raw("COUNT"))
        fines = clean_int(_raw("FINES"))
        arrests = clean_int(_raw("ARRESTS"))
        charges = clean_int(_raw("CHARGES"))

        detection_method = _raw("DETECTION_METHOD") or "Unknown"
        best_detection = _raw("BEST_DETECTION_METHOD") or "Unknown"

        def is_yes(val: str | None) -> bool:
            return val == "Yes"

        record = {
            "id": f"drug-{row_index - 1:05d}",
            "source_file": SOURCE_FILE.name,
            "dataset_key": "police-enforcement-positive-drug-tests",
            "series_year": series_year,
            "period_start": period_start,
            "period_end": period_end,
            "period_label": str(series_year),
            "jurisdiction": jurisdiction,
            "jurisdiction_slug": slugify(jurisdiction),
            "location": location,
            "location_slug": slugify(location),
            "age_group": age_group,
            "age_group_slug": slugify(age_group),
            "metric_label": "Positive drug tests",
            "detection_method": detection_method,
            "detection_method_slug": slugify(detection_method),
            "best_detection_method": best_detection,
            "cannabis_detected": is_yes(_raw("CANNABIS")),
            "amphetamine_detected": is_yes(_raw("AMPHETAMINE")),
            "cocaine_detected": is_yes(_raw("COCAINE")),
            "ecstasy_detected": is_yes(_raw("ECSTASY")),
            "methylamphetamine_detected": is_yes(_raw("METHYLAMPHETAMINE")),
            "other_detected": is_yes(_raw("OTHER")),
            "unknown_detected": is_yes(_raw("UNKNOWN")),
            "no_drugs_detected": is_yes(_raw("NO_DRUGS_DETECTED")),
            "count": count,
            "fines": fines,
            "arrests": arrests,
            "charges": charges,
            "total_actions": count + fines + arrests + charges,
        }
        records.append(record)
        quality["rows_emitted"] += 1

    quality["missing_required_values"] = dict(
        sorted(quality["missing_required_values"].items())
    )
    return records, quality


def aggregate_timeseries(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    yearly: dict[int, dict[str, Any]] = {}
    for record in records:
        year = record["series_year"]
        bucket = yearly.setdefault(
            year,
            {
                "year": year,
                "positive_tests": 0,
                "fines": 0,
                "arrests": 0,
                "charges": 0,
                "total_actions": 0,
            },
        )
        bucket["positive_tests"] += record["count"]
        bucket["fines"] += record["fines"]
        bucket["arrests"] += record["arrests"]
        bucket["charges"] += record["charges"]
        bucket["total_actions"] += record["total_actions"]
    return [yearly[key] for key in sorted(yearly.keys())]


def aggregate_by_jurisdiction(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest_year = max(record["series_year"] for record in records)
    grouped: dict[str, dict[str, Any]] = {}
    for record in records:
        if record["series_year"] != latest_year:
            continue
        bucket = grouped.setdefault(
            record["jurisdiction"],
            {
                "jurisdiction": record["jurisdiction"],
                "positive_tests": 0,
                "fines": 0,
                "arrests": 0,
                "charges": 0,
                "total_actions": 0,
            },
        )
        bucket["positive_tests"] += record["count"]
        bucket["fines"] += record["fines"]
        bucket["arrests"] += record["arrests"]
        bucket["charges"] += record["charges"]
        bucket["total_actions"] += record["total_actions"]
    return sorted(grouped.values(), key=lambda row: -row["total_actions"])


def aggregate_by_jurisdiction_timeseries(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[int, str], int] = defaultdict(int)
    for record in records:
        grouped[(record["series_year"], record["jurisdiction"])] += record["count"]
    rows = [
        {
            "year": year,
            "jurisdiction": jurisdiction,
            "positive_tests": total,
        }
        for (year, jurisdiction), total in grouped.items()
    ]
    return sorted(rows, key=lambda row: (row["year"], row["jurisdiction"]))


def aggregate_substance_trends(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    yearly: dict[int, dict[str, int]] = {}
    for record in records:
        year = record["series_year"]
        bucket = yearly.setdefault(
            year,
            {
                "year": year,
                "cannabis": 0,
                "amphetamine": 0,
                "cocaine": 0,
                "ecstasy": 0,
                "methylamphetamine": 0,
                "other": 0,
            },
        )
        if record["cannabis_detected"]:
            bucket["cannabis"] += 1
        if record["amphetamine_detected"]:
            bucket["amphetamine"] += 1
        if record["cocaine_detected"]:
            bucket["cocaine"] += 1
        if record["ecstasy_detected"]:
            bucket["ecstasy"] += 1
        if record["methylamphetamine_detected"]:
            bucket["methylamphetamine"] += 1
        if record["other_detected"]:
            bucket["other"] += 1
    return [yearly[key] for key in sorted(yearly.keys())]


def aggregate_by_detection_stage(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[int, str], int] = defaultdict(int)
    for record in records:
        grouped[(record["series_year"], record["detection_method"])] += record["count"]
    rows = [
        {
            "year": year,
            "detection_method": method,
            "positive_tests": total,
        }
        for (year, method), total in grouped.items()
    ]
    return sorted(rows, key=lambda row: (row["year"], row["detection_method"]))


def build_summary(
    records: list[dict[str, Any]], quality: dict[str, Any]
) -> dict[str, Any]:
    latest_year = max(record["series_year"] for record in records)
    latest_records = [
        record for record in records if record["series_year"] == latest_year
    ]

    totals = {
        "total_positive_tests": sum(record["count"] for record in records),
        "total_fines": sum(record["fines"] for record in records),
        "total_arrests": sum(record["arrests"] for record in records),
        "total_charges": sum(record["charges"] for record in records),
        "total_actions": sum(record["total_actions"] for record in records),
    }
    latest_totals = {
        "total_positive_tests": sum(record["count"] for record in latest_records),
        "total_fines": sum(record["fines"] for record in latest_records),
        "total_arrests": sum(record["arrests"] for record in latest_records),
        "total_charges": sum(record["charges"] for record in latest_records),
        "total_actions": sum(record["total_actions"] for record in latest_records),
    }

    period_starts = sorted({record["period_start"] for record in records})
    period_ends = sorted({record["period_end"] for record in records})

    return {
        "dataset_key": "police-enforcement-positive-drug-tests",
        "title": "Positive drug tests",
        "description": "Cleaned annual police positive roadside drug test records published by BITRE and the National Road Safety Data Hub.",
        "source": {
            "name": "Bureau of Infrastructure and Transport Research Economics (BITRE)",
            "publication_url": "https://www.bitre.gov.au/publications/2024/road-safety-enforcement-data",
            "dataset_url": "https://datahub.roadsafety.gov.au/safe-systems/safe-road-use/police-enforcement",
            "source_file": SOURCE_FILE.name,
        },
        "schema_version": 1,
        "generated_at": datetime.now(UTC)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z"),
        "record_count": len(records),
        "latest_year": latest_year,
        "period_coverage": {
            "start": period_starts[0],
            "end": period_ends[-1],
            "count": len(period_starts),
        },
        "totals": totals,
        "latest_year_totals": latest_totals,
        "dimensions": {
            "jurisdictions": sorted({record["jurisdiction"] for record in records}),
            "locations": sorted({record["location"] for record in records}),
            "age_groups": sorted({record["age_group"] for record in records}),
            "detection_methods": sorted(
                {record["detection_method"] for record in records}
            ),
        },
        "quality": quality,
        "aggregates": {
            "timeseries": aggregate_timeseries(records),
            "by_jurisdiction": aggregate_by_jurisdiction(records),
            "by_jurisdiction_timeseries": aggregate_by_jurisdiction_timeseries(records),
            "substance_trends": aggregate_substance_trends(records),
            "by_detection_stage": aggregate_by_detection_stage(records),
        },
        "notes": [
            "The source data is annual only; no sub-annual temporal resolution is available.",
            "Substance detection fields (cannabis, amphetamine, cocaine, ecstasy, methylamphetamine, other) are indicator-based (Yes/No/Not applicable) rather than quantitative counts.",
            "Zero values are retained as observed values rather than treated as missing data.",
            "Date values from Excel are converted to ISO 8601 calendar dates for easier downstream use.",
            "Cross-jurisdiction comparisons should be read as indicative because state and territory reporting systems are not fully harmonised.",
        ],
    }


def write_csv(records: list[dict[str, Any]], path: Path) -> None:
    fieldnames = list(records[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    ws = wb.active

    records, quality = build_records(ws)
    records.sort(
        key=lambda record: (
            record["series_year"],
            record["jurisdiction"],
            record["location"],
            record["age_group"],
            record["detection_method"],
        )
    )
    summary = build_summary(records, quality)

    OUTPUT_RECORDS.write_text(json.dumps(records, indent=2), encoding="utf-8")
    OUTPUT_SUMMARY.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    write_csv(records, OUTPUT_CSV)

    print(f"Wrote {len(records)} records to {OUTPUT_RECORDS.name}")
    print(f"Wrote summary to {OUTPUT_SUMMARY.name}")
    print(f"Wrote CSV to {OUTPUT_CSV.name}")


if __name__ == "__main__":
    main()
